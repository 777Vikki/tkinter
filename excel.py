from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from re import sub
from os.path import basename

def compareExcelData(path1, path2):
    fileName1 = basename(path1) + " file"
    fileName2 = basename(path2) + " file"
    excelData1 = excelParsing(path1)
    excelData2 = excelParsing(path2)

    excelList = [["Trans_Id", "Target"], [fileName1, "File Name"], [" ", " "], [" ", " "]]
    for data in excelData1:
        dataList = []
        if(data not in excelData2.keys()):
            print
            dataList.append(data)
            dataList.append(excelData1[data])
            excelList.append(dataList)
    excelList.append([" "," "])
    excelList.append([" "," "])
    excelList.append([fileName2,"File Name"])
    excelList.append([" "," "])
    excelList.append([" "," "])
    for data in excelData2:
        dataList = []
        if(data not in excelData1.keys()):
            dataList.append(data)
            dataList.append(excelData2[data])
            excelList.append(dataList)
    if(len(excelList) != 9):
        xmlToExcel(excelList)
        return len(excelList)
    else:
        return len(excelList)

def xmlToExcel(excelData):
    wb = Workbook()
    sheet = wb.active
    sheet.title = "Localization_Sheet"
    for data in excelData:
        sheet.append(data)
    try:
        refLen = "A1:" + str(chr(len(excelData[0])+64)) +str(len(excelData))
        tab = Table(displayName="Table1", ref=refLen)
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        sheet.add_table(tab)
    except Exception:
        pass
    wb.save("localization.xlsx")

def excelParsing(path):
    wb_obj = load_workbook(path)
    sheet_obj = wb_obj.active
    rows = sheet_obj.max_row
    sheetDataList = {}
    if(rows >= 2):
        for row in range(2, rows + 1):
            sheetId = sheet_obj.cell(row = row, column = 1).value
            if(sheetId != None):
                sheetData = sheet_obj.cell(row = row, column = 5).value
                if(sheetData != None):
                    sheetDataList[sheetId] = sub(' +', ' ', sheetData.strip().replace('--', ''))
                else:
                    sheetDataList[sheetId]= ""
        return sheetDataList